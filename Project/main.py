from os import name
import string
import time
from openpyxl.workbook import workbook
from openpyxl.workbook.workbook import Workbook
import requests
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.utils import find_connectable_ip
from selenium.webdriver.firefox.options import Options
import openpyxl

"""
    Observação: Verificar possibilidade de tranformar num sistema assicrono, desta forma transformando todas as atualizações em tempo real baseado na mudança de cada item em cada pagína.
    Maybe uma possibilidade disso ocorrer também seria listando as pages pela quantidade de page_active apartir do momento que a steam utiliza em sua URL a terminação "p1_name_asc" p2, p3...
    Desta forma aumentando o desempenho da aplicação e otimizando ela como um todo.
"""








######
url = "https://steamcommunity.com/market/search?q=&category_730_ItemSet%5B%5D=any&category_730_ProPlayer%5B%5D=any&category_730_StickerCapsule%5B%5D=any&category_730_TournamentTeam%5B%5D=any&category_730_Weapon%5B%5D=any&category_730_Type%5B%5D=tag_CSGO_Type_Pistol&category_730_Type%5B%5D=tag_CSGO_Type_SMG&category_730_Type%5B%5D=tag_CSGO_Type_Rifle&category_730_Type%5B%5D=tag_CSGO_Type_SniperRifle&category_730_Type%5B%5D=tag_CSGO_Type_Shotgun&category_730_Type%5B%5D=tag_CSGO_Type_Machinegun&category_730_Type%5B%5D=tag_CSGO_Type_Knife&category_730_Type%5B%5D=tag_Type_Hands&appid=730#p1_name_asc"

option = Options()
option.headless = False
driver = webdriver.Firefox(options=option)

driver.get(url)

#Pagina atual até a ultima pagina

pages = int(driver.find_element_by_xpath('//*[@id="searchResults_links"]/span[7]').text)
pages = pages - (pages - 2)
page_active = int(driver.find_element_by_xpath('//*[@id="searchResults_links"]/*[@class="market_paging_pagelink active"]').text)
time.sleep(5)
type = []
name = []
stattrak = []
souvenir = []
floats = []
value = []

#Passar as páginas e pegar o conteúdo
print(page_active, pages)
skin_counter = 0
while page_active <= pages:
    print('-='*15, page_active, '=-'*15)
    
    

    #Conteúdo da página
    i = 0
    while i < 10:
        str = '//*[@id="result_{}_name"]'.format(i)
        weapon = driver.find_element_by_xpath('//div[@class="market_listing_item_name_block"]{}'.format(str)).text

        #StartTrak
        a = weapon[weapon.find(' (★'): weapon.find(') ') + 2]
        if '(★)' in weapon:
            weapon_stattrak = '★'
            weapon = weapon.replace(a, '')
        elif '(★ StatTrak™)' in weapon:
                weapon_stattrak = '★ StatTrak™'
                weapon = weapon.replace(a, '')      
        elif '(StatTrak™)' in weapon:
            weapon_stattrak = 'StatTrak™'
            weapon = weapon.replace('(StatTrak™)', '')
        else:
            weapon_stattrak = 'No'


        #Souvenir
        if ('Lembrança') in weapon:
            weapon_souvenir = 'Solvenir'
            weapon = weapon.replace(' (Lembrança)', '')
        else:
            weapon_souvenir = 'No'     
        if '|' in weapon:
            #Name
            weapon_name = weapon[weapon.find('| ') + 2: weapon.find(' (')].strip()

            #Type
            weapon_type = weapon[:weapon.find('|')].strip()

            #Float
            weapon_float_name = weapon[weapon.find(' (') + 2: len(weapon) - 1]
        else:
            #Name
            weapon_name = weapon_type = weapon[:weapon.find(' (')]

            #Float
            weapon_float_name = 'All'
        

        #Value
        val = '//*[@id="result_{}"]/div[1]/div[2]/span[1]/span[1]'.format(i)
        weapon_value = driver.find_element_by_xpath(val).text
        weapon_value = float(weapon_value[1:len(weapon_value) - 4:].replace(',', ''))

        print(weapon_type, '|',weapon_name, '({})'.format(weapon_stattrak), '({})'.format(weapon_float_name), '- Cost: USD ${:.2f}'.format(weapon_value))
        
        type.append(weapon_type)
        name.append(weapon_name)
        stattrak.append(weapon_stattrak)
        souvenir.append(weapon_souvenir)
        floats.append(weapon_float_name)
        value.append(weapon_value)
        i += 1
    skin_counter += i
    
    #Steam Access Delay 
    last_page = page_active
    driver.find_element_by_xpath('//*[@id="searchResults_btn_next"]').click()
    time.sleep(5)
    page_active = int(driver.find_element_by_xpath('//*[@id="searchResults_links"]/*[@class="market_paging_pagelink active"]').text)
    t = 0
    while page_active == last_page:
        time.sleep(1)
        page_active = int(driver.find_element_by_xpath('//*[@id="searchResults_links"]/*[@class="market_paging_pagelink active"]').text)
        if t == 6:
            time.sleep(60)
            t = 0
        t += 1

all_items_list = {'type': type, 'name': name, 'stattrack': stattrak, 'souvenir': souvenir, 'floats': floats, 'value': value}
tabless = pd.DataFrame(data=all_items_list)
print(tabless)
wb =  openpyxl.load_workbook(r'C:\Users\Pentest\Documents\DB Items.xlsx')
sheet = wb.sheetnames
print(sheet)
for worksheets in wb.sheetnames:

    worksheet = wb[worksheets]
    i = 0
    for k, v in all_items_list.items():
        i += 1
        print(i, '-', k, v)


wb.save(r'C:\Users\Pentest\Documents\DB Items.xlsx')
time.sleep(5)
driver.quit()

