import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles.fills import Fill
from openpyxl.workbook import workbook
from openpyxl.workbook.workbook import Workbook
from selenium import webdriver
from selenium.webdriver.common.utils import find_connectable_ip
from selenium.webdriver.firefox.options import Options
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.styles import colors
"""
option = Options()
option.headless = False
driver = webdriver.Firefox(options=option)

driver.get(url)

#Pagina atual até a ultima pagina

pages = int(driver.find_element_by_xpath('//*[@id="searchResults_links"]/span[7]').text)
pages = pages - (pages - 2)
page_active = int(driver.find_element_by_xpath('//*[@id="searchResults_links"]/*[@class="market_paging_pagelink active"]').text)
time.sleep(7)
types_list = []
name_list = []
stattrak_list = []
souvenir_list = []
floats_list = []
value_list = []

#Passar as páginas e pegar o conteúdo
print(page_active, pages)
skin_counter = 0
while page_active <= pages:
    print('-='*15, page_active, '=-'*15)
    
    

    #Conteúdo da página
    i = 0
    while i < 10:
        str = '//*[@id="result_{}_name"]'.format(i)
        weapon = driver.find_element_by_xpath('//div[@class="market_listing_item_name_block"]{}'.format(str)).text

        #StartTrak
        a = weapon[weapon.find(' (★'): weapon.find(') ') + 2]
        if '(★)' in weapon:
            weapon_stattrak = '★'
            weapon = weapon.replace(a, '')
        elif '(★ StatTrak™)' in weapon:
                weapon_stattrak = '★ StatTrak™'
                weapon = weapon.replace(a, '')      
        elif '(StatTrak™)' in weapon:
            weapon_stattrak = 'StatTrak™'
            weapon = weapon.replace('(StatTrak™)', '')
        else:
            weapon_stattrak = 'No'


        #Souvenir
        if ('Lembrança') in weapon:
            weapon_souvenir = 'Solvenir'
            weapon = weapon.replace(' (Lembrança)', '')
        else:
            weapon_souvenir = 'No'     
        if '|' in weapon:
            #Name
            weapon_name = weapon[weapon.find('| ') + 2: weapon.find(' (')].strip()

            #Type
            weapon_type = weapon[:weapon.find('|')].strip()

            #Float
            weapon_float_name = weapon[weapon.find(' (') + 2: len(weapon) - 1]
        else:
            #Name
            weapon_name = weapon_type = weapon[:weapon.find(' (')]

            #Float
            weapon_float_name = 'All'
        

        #Value
        val = '//*[@id="result_{}"]/div[1]/div[2]/span[1]/span[1]'.format(i)
        weapon_value = driver.find_element_by_xpath(val).text
        weapon_value = float(weapon_value[1:len(weapon_value) - 4:].replace(',', ''))

        print(weapon_type, '|',weapon_name, '({})'.format(weapon_stattrak), '({})'.format(weapon_float_name), '- Cost: USD ${:.2f}'.format(weapon_value))
        
        types_list.append(weapon_type)
        name_list.append(weapon_name)
        stattrak_list.append(weapon_stattrak)
        souvenir_list.append(weapon_souvenir)
        floats_list.append(weapon_float_name)
        value_list.append(weapon_value)
        i += 1
    skin_counter += i
    
    #Steam Access Delay 
    last_page = page_active
    driver.find_element_by_xpath('//*[@id="searchResults_btn_next"]').click()
    time.sleep(5)
    page_active = int(driver.find_element_by_xpath('//*[@id="searchResults_links"]/*[@class="market_paging_pagelink active"]').text)
    t = 0
    while page_active == last_page:
        time.sleep(1)
        page_active = int(driver.find_element_by_xpath('//*[@id="searchResults_links"]/*[@class="market_paging_pagelink active"]').text)
        if t == 6:
            time.sleep(60)
            t = 0
        t += 1

all_items_list = {'type': types_list, 'name': name_list, 'stattrack': stattrak_list, 'souvenir': souvenir_list, 'floats': floats_list, 'value': value_list}
tabless = pd.DataFrame(data=all_items_list)
print(tabless)
wb = load_workbook(r'C:\Users\Pentest\Documents\DB Items.xlsx')






for worksheets in wb.sheetnames:
    dist_top = 4
    dist_left = 2
    worksheet = wb[worksheets]
    print(worksheet)
    for i in range(1, len(types_list) + 50):
        row = worksheet.row_dimensions[i]
        row.fill = PatternFill("solid", fgColor="262626")
    for i in range(dist_top, len(types_list) + dist_top):
        
        
        
        #Cells Values
        #TRANSFORMAR EM OBJETO PELO AMOR DE GOD, ME NOTA SENPAI
        type_cell = worksheet.cell(row=i, column=1+dist_left, value=types_list[i-dist_top])
        type_cell.font = Font(name='Segoe UI', size=12)
        type_cell.fill = PatternFill("solid", fgColor="262626")

        name_cell = worksheet.cell(row=i, column=2+dist_left, value=name_list[i-dist_top])
        name_cell.font = Font(name='Segoe UI', size=12)
        name_cell.fill = PatternFill("solid", fgColor="262626")

        stattrak_cell = worksheet.cell(row=i, column=3+dist_left, value=stattrak_list[i-dist_top])
        stattrak_cell.font = Font(name='Segoe UI', size=12)
        stattrak_cell.fill = PatternFill("solid", fgColor="262626")

        souvenir_cell = worksheet.cell(row=i, column=4+dist_left, value=souvenir_list[i-dist_top])
        souvenir_cell.font = Font(name='Segoe UI', size=12)
        souvenir_cell.fill = PatternFill("solid", fgColor="262626")

        dist_cell = worksheet.cell(row=i, column=5+dist_left, value=floats_list[i-dist_top])
        dist_cell.font = Font(name='Segoe UI', size=12)
        dist_cell.fill = PatternFill("solid", fgColor="262626")

        value_cell = worksheet.cell(row=i, column=6+dist_left, value=value_list[i-dist_top])
        value_cell.font = Font(name='Segoe UI', size=12)
        value_cell.fill = PatternFill("solid", fgColor="262626")

    #Titles    
    worksheet.cell(row=dist_top-1,column=1+dist_left,value="Type").font = Font(name='Segoe UI', size=14)
    worksheet.cell(row=dist_top-1,column=2+dist_left,value="Name").font = Font(name='Segoe UI', size=14)
    worksheet.cell(row=dist_top-1,column=3+dist_left,value="Stattrak").font = Font(name='Segoe UI', size=14)
    worksheet.cell(row=dist_top-1,column=4+dist_left,value="Souvenir").font = Font(name='Segoe UI', size=14)
    worksheet.cell(row=dist_top-1,column=5+dist_left,value="Float Name").font = Font(name='Segoe UI', size=14)
    worksheet.cell(row=dist_top-1,column=6+dist_left,value="Value ($)").font = Font(name='Segoe UI', size=14)




wb.save(r'C:\Users\Pentest\Documents\DB Items.xlsx')
time.sleep(5)
driver.quit()
"""